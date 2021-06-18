from openpyxl import Workbook
from bs4 import BeautifulSoup
import requests as rq
from openpyxl.chart import LineChart,Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import datetime
import time


def alert_bot(subject,content,to,url,file_name):
    email_user = 'stonkman0@gmail.com'
    email_password = 'zmkatlhdwlxuxnsm'

    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = to
    msg['Subject'] = subject

    body = content + '\n\n'+ 'Link: ' + url + '\n\n'
    msg.attach(MIMEText(body,'plain'))

    filename=file_name
    attachment  = open(filename,'rb')

    part = MIMEBase('application','octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition',"attachment; filename= "+filename)

    msg.attach(part)
    text = msg.as_string()
    server = smtplib.SMTP('smtp.gmail.com',587)
    server.starttls()
    server.login(email_user,email_password)

    server.sendmail(email_user,to,text)
    server.quit()


def realtime_fetch(url):
    
    time_now_list=[]
    list = []
    name_in=input("Please provide a backup file name: ")
    tym=input("Enter time(HH:MM): ")

    cond = ''
    while cond != 'N':

        cond = input("\nDo you want the file as a mail?(Y/N): ").upper()

        if cond == 'Y':
            to = input("\nEnter your mail id: ")
            mail_val = 1
            break

        elif cond == 'N':
            mail_val = 0
            print('\nFile will be saved!')

        else:
            print("\nInvalid entry ")

    print("Fetching information in realtime.You can minimize this window do not close. ")

    while time.strftime('%I:%M')!=tym:

        html = rq.get(url).text
        soup = BeautifulSoup(html, 'lxml')
        name_1 = soup.find('h1', class_='D(ib) Fz(18px)')

        if name_1!=None:
            name=name_1.text.replace(' ', '')
        elif name_1==None:
            name=name_in

        time_now=time.strftime('%I:%M:%S')
        time_now_list.append(time_now)

        rltp = soup.find('span', class_='Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)')

        if rltp!=None:
            list.append(rltp.text.replace(',', ''))
        elif rltp==None:
            list.append('0')

        time.sleep(3)

    print('1:')
    new_list=[]
    i=0
    while i<len(list):
        x=list[i].replace(',','')
        new_list.append(x)
        i+=1

    print('2:')
    i=0
    formatted_list=[]
    while i<len(new_list):
        y=float(new_list[i])
        formatted_list.append(y)
        i+=1

    print('3:')

    sum=0
    i=0
    while i<len(formatted_list):
        sum+=formatted_list[i]
        i+=1

    print('4:')
    avg=sum/len(formatted_list)

    wb = Workbook()
    sheet = wb.active
    sheet.title = 'DayPrice'
    headings = ['Time',f'{name}']
    sheet.append(headings)
    print('5:')

    for row in range(2,len(time_now_list)+1):
        sheet['A'+f'{row}']=time_now_list[row-2]

    for row in range(2,len(formatted_list)+1):
        sheet['B'+f'{row}']=formatted_list[row-2]

    for row in range(2,len(formatted_list)+1):
        if sheet['B'+f'{row}'].value > avg:
            sheet['B'+f'{row}'].fill=PatternFill(start_color='00339966',
                                                 end_color='00339966',
                                                 fill_type='solid')

        elif sheet['B'+f'{row}'].value < avg:
            sheet['B' + f'{row}'].fill = PatternFill(start_color='00FF8080',
                                                     end_color='00FF8080',
                                                     fill_type='solid')

        elif sheet['B'+f'{row}'].value == avg:
            sheet['B' + f'{row}'].fill = PatternFill(start_color='00FFCC00',
                                                     end_color='00FFCC00',
                                                     fill_type='solid')

    print('6:')
    refVal= Reference(sheet,
                      max_row=len(formatted_list)+1,
                      max_col=2,
                      min_row=2,
                      min_col=2)

    print('7:')
    chart=LineChart()
    chart.add_data(refVal)
    sheet.add_chart(chart,'D2')
    print('8:')
    filename=name+'.xlsx'
    wb.save(filename)

    if mail_val==1:
        print('mail sent!')
        alert_bot('Stock details',f'Todays stock details of {name}',to,url,filename)


url=input('Please provide the page url: ')
realtime_fetch(url)




# antoanson6@gmail.com

# https://in.finance.yahoo.com/quote/NVDA/history?p=NVDA
# https://in.finance.yahoo.com/quote/AMD/history?p=AMD
# https://in.finance.yahoo.com/quote/MSFT/history?p=MSFT

# python rltprice.py
